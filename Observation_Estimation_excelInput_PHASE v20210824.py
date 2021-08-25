#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Fri Aug 10 15:21:02 2018

@author: ChrisLin
"""

# === Section 0: basic tools ===
import matplotlib.pyplot as plot
import numpy as np
import os
import time
import openpyxl
import math

servo_choice = 2 #0: pre_servo with calibration 1: post_servo with calibration 2: pre_servo without calibration
pole = ['X', 'Y']

def import_data():
    exl = 'moon_calibration.xlsx'
    wb = openpyxl.load_workbook(exl)
    ymds = wb.get_sheet_names()
    in_load_cal = []
    in_cbx1 = in_cbx2 = in_cbx3 = in_cby1 = in_cby2 = in_cby3 = []
    gain = np.empty([len(ymds), 2, 7]) #[day][pole][Rx]
    
    if servo_choice == 0 or 2:
        servo = '_preservo'
        for i in range(len(ymds)):
            ws = wb.get_sheet_by_name(ymds[i])
            in_load_cal.append(ws['B2'].value)
            gain[i][0] = [1, 1, 1, 1, 1, 1, 1]
        for i in range(len(ymds)):
            ws = wb.get_sheet_by_name(ymds[i])
            in_load_cal.append(ws['C2'].value) 
            gain[i][1] = [1, 1, 1, 1, 1, 1, 1]
    else:  
        servo = '_postservoCor'
        for i in range(len(ymds)):
            ws = wb.get_sheet_by_name(ymds[i])
            in_load_cal.append(ws['B3'].value)
            j = 0
            for row in ws['F2':'L2']:
                for cell in row:        
                    gain[i][0][j] = cell.value
                    j += 1
        for i in range(len(ymds)):
            ws = wb.get_sheet_by_name(ymds[i])
            in_load_cal.append(ws['C3'].value) 
            j = 0
            for row in ws['F3':'L3']:
                for cell in row:        
                    gain[i][1][j] = cell.value
                    j += 1
    ws = wb.get_sheet_by_name(ymds[0])
    for i in range(5, ws.max_row+1):
        if ws.cell(column=2, row=i).value:
            in_cbx1.append(ws.cell(column=2, row=i).value)
        if ws.cell(column=3, row=i).value:
            in_cby1.append(ws.cell(column=3, row=i).value)
    ws = wb.get_sheet_by_name(ymds[1])
    for i in range(5, ws.max_row+1):
        if ws.cell(column=2, row=i).value:
            in_cbx2.append(ws.cell(column=2, row=i).value)
        if ws.cell(column=3, row=i).value:
            in_cby2.append(ws.cell(column=3, row=i).value)
    ws = wb.get_sheet_by_name(ymds[2])
    for i in range(5, ws.max_row+1):
        if ws.cell(column=2, row=i).value:
            in_cbx3.append(ws.cell(column=2, row=i).value)
        if ws.cell(column=3, row=i).value:
            in_cby3.append(ws.cell(column=3, row=i).value)
    return ymds, in_load_cal, servo, in_cbx1, in_cby1, in_cbx2, in_cby2, in_cbx3, in_cby3, gain

ymds, load_cal, servo, cbx1, cby1, cbx2, cby2, cbx3, cby3, gain = import_data()
print(servo)

def vis_plot(j, a, p, m, n, path, chan_low, chan_up, show_cal_amp, BLtype, bin_chan_low, bin_chan_up, bin_show_cal_amp):
    vb1 = vb2 = vb3 = vb4 = []
    fn1 = 'multi_spectrum_data_Tfade_full_moon_phase.xlsx'
    wb1 = openpyxl.load_workbook(fn1)
    ws1 = wb1.get_sheet_by_name('Sheet')
    for i in range(1, ws1.max_column+1):
        vb1.append(ws1.cell(column=i, row=1+4*j).value)
        vb2.append(ws1.cell(column=i, row=2+4*j).value)
        vb3.append(ws1.cell(column=i, row=3+4*j).value)
        vb4.append(ws1.cell(column=i, row=4+4*j).value)
    for i in range(1, ws1.max_column+1):
        plot.scatter(i * 10.0 / 1024 + 94 - 10/2, vb1[i-1], color = 'b')
        plot.scatter(i * 10.0 / 1024 + 94 - 10/2, vb2[i-1], color = 'g')
        plot.scatter(i * 10.0 / 1024 + 94 - 10/2, vb3[i-1], color = 'gold')
        plot.scatter(i * 10.0 / 1024 + 94 - 10/2, vb4[i-1], color = 'r')
    
    #plot.scatter(chan * 5 / 1024 + 89, show_cal_amp[a][p][0][j], s = 2)
    plot.scatter(bin_chan_low*bin_n * 5 / 1024 + 89, bin_show_cal_amp[a][p][0][j], s = 10, color = 'darkmagenta')
    plot.scatter(bin_chan_up*bin_n * 5 / 1024 + 94, bin_show_cal_amp[a][p][1][j], s = 10, color = 'magenta')
    plot.title('%s_%s%s_BL: %d-%d)' % (ymds[a], pole[p], servo, m, n))
    plot.ylabel('Cal_Visibility (Jy)')
    plot.grid('True')
    plot.savefig(path + '/%s_%s_%s_Flux_BL %d-%d.png' % (ymds[a], pole[p], servo, m, n))
    
    plot.show()
        


file_name = 'observation_%s' % (time.strftime("%Y-%m-%d %H-%M-%S", time.localtime()))
os.mkdir(file_name)


# 1-5. load calibrated data (by using loadh5.py)
from loadh5 import *

if servo_choice == 2:
    cal1, cal2, cal3, cal4, cal5, cal6 = [ldrawvis(load_cal[i]) for i in range(6)]
else:
    cal1, cal2, cal3, cal4, cal5, cal6 = [ldcalvis(load_cal[i]) for i in range(6)]

calx = [cal1] + [cal2] + [cal3]
caly = [cal4] + [cal5] + [cal6]

cals = np.empty([len(ymds), 2, 2, 21, 1024, 2], dtype = complex) # day, pole, usb/lsb, BL, chan, fix0
for h in range(len(ymds)):
    for i in range(2):
        for j in range(21):
            for k in range(1024):
                cals[h][0][i][j][k][0] = calx[h][1][i][j][k][1]
                cals[h][1][i][j][k][0] = caly[h][1][i][j][k][1]
            

# === Section 2: other setting ===
# 2-1. time setting
import datetime
t = datetime.datetime.now()

# 2-2. channel setting
chan_min_max = [0, 1024]
chan_nub = chan_min_max[1] - chan_min_max[0] - 1

# 2-3. plotting max setting ([0: no max, 1: cut off extreme values], [max cutting value])
cut = [1, 8000]

# 2-4. error bar setting (0: no error bars, 1: 1 sigma error bars)
err = 1


# === Section 3: Qualify data ===
cutbase = np.empty([len(ymds)], dtype = int)

cb1, cb2, cb3 = [cbx1, cby1], [cbx2, cby2], [cbx3, cby3]
#cb4, cb5, cb6 = [cbx4, cby4], [cbx5, cby5], [cbx6, cby6]
cutbase = [cb1] + [cb2] + [cb3] # + [cb4] + [cb5] + [cb6]
cutbase_leng = [[len(cbx1), len(cby1)], [len(cbx2), len(cby2)], [len(cbx3), len(cby3)]]


# === Section 4: Calculation setting ===
# 4-1. visibility magnitude, standard diviation, baseline length, channel
avg_flux = np.empty([len(ymds), 2, 21])
sd_flux = np.empty([len(ymds), 2, 21])
base_length = np.empty([len(ymds), 21])
chan_low = np.empty([chan_min_max[1]])
chan_up = np.empty([chan_min_max[1]])
show_cal_amp = np.empty([len(ymds), 2, 2, 21, chan_min_max[1]], dtype = float)


bin_n = 8
bin_chan_max = chan_min_max[1]/bin_n
bin_chan_low = np.empty(bin_chan_max)
bin_chan_up = np.empty(bin_chan_max)
bin_show_cal_amp = np.empty([len(ymds), 2, 2, 21, bin_chan_max], dtype = float)

BLtype = np.empty([21])
# 4-6. output path setting
path = '%s/plot_BL_%dymds_%s_to_%s_servoCor_%s' % (file_name, len(ymds), ymds[0], ymds[len(ymds)-1], servo)
if cut[0] == 1:
    path += '_2ndCal'
    if err == 1:
        path += '_errbar'

if not os.path.exists(path):
    os.mkdir(path)
    
# === Section 5: Calculation ===
# time record
for a in range(len(ymds)):
    f = open(path + '/flux_baselength_day' + str(a+1) + '.txt', "a+")
    f.write(str(t) + '\n' + 'file: ' + str(load_cal[a]) + '\n')
    f.close()    
    # calculation
    for p in range(2):
        for j in range(21):
            # 21 baseline to Rx0-6
            if j < 6:
                m = 0
                n = j + 1
            elif j < 11:
                m = 1
                n = j - 4
            elif j < 15:
                m = 2
                n = j - 8
            elif j < 18:
                m = 3
                n = j - 11
            elif j < 20:
                m = 4
                n = j - 13
            else:
                m = 5
                n = 6
            base1 = m
            base2 = n
            # plotting not aligned
            if m == 0:
                base_length[a][j] = 1.4 - 0.24 + 0.1*a + j*0.008
                BLtype[j] = 0
            elif n - m == 2 or n - m == 4:
                base_length[a][j] = 1.4 * (3 ** 0.5) - 0.2 + 0.05*a + j*0.007
                BLtype[j] = 1
            elif n - m == 3:
                base_length[a][j] = 2.8 - 0.02 + 0.02*a + j*0.001
                BLtype[j] = 2
            else:
                base_length[a][j] = 1.4 - 0.24 + 0.1*a + j*0.008
                BLtype[j] = 0
            # setting
            total_flux = total_sqrt_flux = 0
            # averaging visibility magnitude
            for i in range(chan_min_max[0], chan_min_max[1]):
                z = cals[a][p][0][j][i][0]
                z_real = z.real
                z_image = z.imag
                show_cal_amp[a][p][0][j][chan_min_max[1]-1-i] = math.atan(z_real/z_image)
                total_flux += abs(cals[a][p][0][j][i][0])
                total_sqrt_flux += (abs(cals[a][p][0][j][i][0]))**2 
                chan_low[i] = i
                
            for i in range(chan_min_max[0], chan_min_max[1]):
                z = cals[a][p][1][j][i][0]
                z_real = z.real
                z_image = z.imag
                show_cal_amp[a][p][1][j][i] = math.atan(z_real/z_image)
                total_flux += abs(cals[a][p][1][j][i][0])
                total_sqrt_flux += (abs(cals[a][p][1][j][i][0]))**2 
                chan_up[i] = i
            
            bin_sum = bin_count = 0
            for i in range(chan_min_max[0], chan_min_max[1]):
                if bin_count < bin_n:
                    z = cals[a][p][0][j][i][0]
                    z_real = z.real
                    z_image = z.imag
                    bin_sum += math.atan(z_image/z_real)
                else:
                    bin_count = 0
                    bin_show_cal_amp[a][p][0][j][(chan_min_max[1]-1-i)/bin_n] = bin_sum/bin_n
                    z = cals[a][p][0][j][i][0]
                    z_real = z.real
                    z_image = z.imag
                    bin_sum = math.atan(z_image/z_real)
                    bin_chan_low[i/bin_n] = i/bin_n
                bin_count += 1
                
            bin_sum = bin_count = 0
            for i in range(chan_min_max[0], chan_min_max[1]):
                if bin_count < bin_n:
                    z = cals[a][p][1][j][i][0]
                    z_real = z.real
                    z_image = z.imag
                    bin_sum += math.atan(z_image/z_real)
                else:
                    bin_count = 0
                    bin_show_cal_amp[a][p][1][j][i/bin_n] = bin_sum/bin_n
                    z = cals[a][p][1][j][i][0]
                    z_real = z.real
                    z_image = z.imag
                    bin_sum = math.atan(z_image/z_real)
                    bin_chan_up[i/bin_n] = i/bin_n
                bin_count += 1
                

                
            avg_flux[a][p][j] = total_flux / chan_nub
            sd_flux[a][p][j] = (total_sqrt_flux / chan_nub - avg_flux[a][p][j]**2) ** (0.5)
            if cut[0] == 1:
                for k in range(cutbase_leng[a][p]):
                    if j == cutbase[a][p][k]:
                        avg_flux[a][p][j] = -100
                        sd_flux[a][p][j] = 0
            #gain correction        
            avg_flux[a][p][j] = avg_flux[a][p][j] * ((gain[a][p][m] * gain[a][p][n]) ** 0.5)
            sd_flux[a][p][j] = sd_flux[a][p][j] * ((gain[a][p][m] * gain[a][p][n]) ** 0.5)
            
            # document record
            f = open(path + '/flux_baselength_day' + str(a+1) + 'pole' + pole[p] + '.txt', "a+")
            f.write('[base1]:' + str(m) + '  [base2]:' + str(n) + '  [base_length]:' + str("%.2f" % base_length[a][j]) + '  [avg_flux]:' + str("%.2f" % avg_flux[a][p][j]) + '  [sd]:' + str("%.2f" % sd_flux[a][p][j]) + '\n')
            f.close()
            
            vis_plot(j, a, p, m, n, path, chan_low, chan_up, show_cal_amp, BLtype, bin_chan_low, bin_chan_up, bin_show_cal_amp)


# === Section 6: Plotting ===
# 6-1. error bars
for p in range(2):
    for i in range(len(ymds)):
        plot.errorbar(base_length[i], avg_flux[i][p], yerr = sd_flux[i][p], fmt = 'o', capsize = 4)
    
    # 6-2. whole plotting
    if cut[0] == 1:    
        plot.ylim(0, cut[1])
    plot.legend(ymds)
    plot.title('Visibility vs baselengthes (%d ymds) %spol_%s' % (len(ymds), pole[p], servo))
    plot.ylabel('Visibility Magnitude (Jy)')
    plot.grid('True')
    plot.savefig(path + '/Visibility_baselength' + pole[p] + '.png')
    plot.show()